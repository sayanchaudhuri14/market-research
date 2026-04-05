# grid_search_backtest.py
# ─────────────────────────────────────────────────────────────────────────────
# Grid search over:
#   STOP_LOSS_PCT    : [0.20, 0.30, 0.40]
#   PROFIT_TARGET_PCT: [0.30, 0.40, 0.50, 0.60]
#   FIXED            : [5, 10]
#   MAX_LOTS         : [20, 50, 100]
#   DTE0_MAX_LOTS    : [20, 50, 100]  (always ≤ MAX_LOTS)
#
# SIGNAL_MODE is fixed to BEARISH (no edge for BULLISH in real data).
#
# Strategy:
#   • For each (SL, TP) pair → run simulate_trade on all tradeable days once
#   • Cache the raw trade results per (SL, TP) — no re-downloading bhav copies
#   • For each (FIXED, MAX_LOTS, DTE0) on cached results → compute capital metrics
#   • Total combinations: 3 × 4 × 2 × 6 = 144
#
# Drop this file next to your notebook and run:
#   python grid_search_backtest.py
#
# Output: grid_search_results.xlsx  (in OUTPUT_DIR)
# ─────────────────────────────────────────────────────────────────────────────

import numpy as np
import pandas as pd
import pickle, requests, zipfile, io, time as _time
import itertools, warnings
from pathlib import Path
from datetime import date, timedelta, datetime
from datetime import time as dtime
from scipy.stats import norm as _norm
from scipy.optimize import newton

warnings.filterwarnings('ignore')

# ── Fixed params ──────────────────────────────────────────────────────────────
LOT_SIZE       = 75
STRIKE_STEP    = 50
STRIKES_OTM    = 1
RISK_FREE_RATE = 0.065
BROKERAGE      = 80
BACKTEST_DAYS  = 252 * 3
BASE_RATE      = 54.5
SIGNAL_MODE    = 'Bearish'

# ── Grid to sweep ─────────────────────────────────────────────────────────────
SL_GRID   = [0.20, 0.30, 0.40]
TP_GRID   = [0.30, 0.40, 0.50, 0.60]
FIXED_GRID     = [5, 10]
MAX_LOTS_GRID  = [20, 50, 100]
DTE0_LOTS_GRID = [20, 50, 100]   # filtered to ≤ MAX_LOTS below

EXPIRY_CHANGE_DATE = date(2025, 9, 2)

EVENT_DAYS = {
    date(2024, 6, 4), date(2024, 6, 5),
    date(2024, 2, 1), date(2025, 2, 1), date(2026, 2, 1),
    date(2024, 4, 5),  date(2024, 6, 7),  date(2024, 8, 8),
    date(2024, 10, 9), date(2024, 12, 6),
    date(2025, 2, 7),  date(2025, 4, 9),  date(2025, 6, 6),
    date(2025, 8, 6),  date(2025, 10, 8), date(2025, 12, 5),
    date(2026, 2, 6),  date(2026, 4, 3),
    date(2023, 12, 3), date(2024, 11, 23), date(2025, 2, 8),
}

NSE_HOLIDAYS = {
    date(2023,1,26), date(2023,3,7),  date(2023,4,4),  date(2023,4,7),
    date(2023,4,14), date(2023,5,1),  date(2023,6,28), date(2023,8,15),
    date(2023,9,19), date(2023,10,2), date(2023,10,24),date(2023,11,14),
    date(2023,11,27),date(2023,12,25),
    date(2024,1,22), date(2024,1,26), date(2024,3,25), date(2024,4,9),
    date(2024,4,11), date(2024,4,14), date(2024,4,17), date(2024,4,21),
    date(2024,5,1),  date(2024,5,23), date(2024,6,17), date(2024,7,17),
    date(2024,8,15), date(2024,10,2), date(2024,10,14),date(2024,11,1),
    date(2024,11,15),date(2024,12,25),
    date(2025,2,26), date(2025,3,14), date(2025,3,31), date(2025,4,10),
    date(2025,4,14), date(2025,4,18), date(2025,5,1),  date(2025,6,6),
    date(2025,7,29), date(2025,8,15), date(2025,8,27), date(2025,10,2),
    date(2025,10,20),date(2025,10,21),date(2025,11,5), date(2025,12,25),
    date(2026,1,26), date(2026,3,20), date(2026,4,3),  date(2026,4,14),
    date(2026,5,1),  date(2026,8,15), date(2026,10,2), date(2026,11,9),
    date(2026,12,25),
}

# ── Paths (mirrors notebook) ──────────────────────────────────────────────────
BASE         = Path.cwd() if (Path.cwd()/'v2'/'v2_aligned_dataset.csv').exists() else Path.cwd().parent
ALIGNED_CSV  = BASE / 'v2' / 'v2_aligned_dataset.csv'
MINUTE_CACHE = BASE / 'v2' / 'kite_minute_cache'
SIGNALS_CSV  = BASE / 'v2' / 'v2_reliable_signals.csv'
BHAV_CACHE   = BASE / 'backtesting_true_data' / 'bhav_cache'
OUTPUT_DIR   = BASE / 'backtesting_true_data' / 'backtest_outputs'
BHAV_CACHE.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════════
# 1. LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════════
print('Loading aligned dataset...')
aligned = pd.read_csv(ALIGNED_CSV, parse_dates=['india_date'])
aligned = aligned.sort_values('india_date').reset_index(drop=True)
aligned['VIX_INDIA_level'] = aligned['VIX_INDIA_level'].ffill().bfill()
backtest_df = aligned.tail(BACKTEST_DAYS).copy().reset_index(drop=True)
start_date  = backtest_df['india_date'].iloc[0].date()
end_date    = backtest_df['india_date'].iloc[-1].date()

reliable    = pd.read_csv(SIGNALS_CSV)
top_bearish = (reliable[reliable['P_Down'] > BASE_RATE]
               .sort_values('Edge_pp', ascending=False).head(3).reset_index(drop=True))

print(f'Backtest period: {start_date} → {end_date}  ({len(backtest_df)} days)')
print(f'Top BEARISH combos: {list(top_bearish["Signal"])}')

print('Loading NIFTY spot minute cache...')
all_chunks = []
for pkl_path in sorted(MINUTE_CACHE.glob('minute_256265_*.pkl')):
    with open(pkl_path, 'rb') as f:
        chunk = pickle.load(f)
    chunk.index = pd.to_datetime(chunk.index)
    if chunk.index.tzinfo is None:
        chunk.index = chunk.index.tz_localize('Asia/Kolkata')
    lo = (pd.Timestamp(start_date) - pd.Timedelta(days=1)).tz_localize('Asia/Kolkata')
    hi = (pd.Timestamp(end_date)   + pd.Timedelta(days=1)).tz_localize('Asia/Kolkata')
    mask = (chunk.index >= lo) & (chunk.index <= hi)
    if mask.sum() > 0:
        all_chunks.append(chunk[mask])

if all_chunks:
    minute_all     = pd.concat(all_chunks).sort_index()
    _trading_dates = set(minute_all.index.normalize().map(lambda ts: ts.date()))
    print(f'Minute data: {len(minute_all):,} rows')
else:
    minute_all = pd.DataFrame()
    _trading_dates = set()
    print('WARNING: No minute cache found.')

# ═══════════════════════════════════════════════════════════════════════════════
# 2. HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
_nse_session = None

def _get_nse_session():
    global _nse_session
    if _nse_session is not None:
        return _nse_session
    s = requests.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Referer': 'https://www.nseindia.com/',
    })
    for warmup_url in ['https://www.nseindia.com',
                        'https://www.nseindia.com/market-data/futures-and-options-market-data']:
        try: s.get(warmup_url, timeout=12); _time.sleep(0.5)
        except: pass
    _nse_session = s
    return s

def _parse_bhav_zip(raw_bytes):
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
            df = pd.read_csv(z.open(z.namelist()[0]))
    except: return None
    df.columns = df.columns.str.strip()
    if 'TckrSymb' in df.columns:
        df = df.rename(columns={'TckrSymb':'symbol','XpryDt':'expiry_raw',
                                 'StrkPric':'strike','OptnTp':'opt_type','OpnPric':'open'})
        df['expiry'] = pd.to_datetime(df['expiry_raw'], errors='coerce').dt.date
    elif 'SYMBOL' in df.columns:
        df = df.rename(columns={'SYMBOL':'symbol','EXPIRY_DT':'expiry_raw',
                                 'STRIKE_PR':'strike','OPTION_TYP':'opt_type','OPEN':'open'})
        df['expiry'] = pd.to_datetime(df['expiry_raw'], dayfirst=True, errors='coerce').dt.date
    else: return None
    df = df[df['symbol'].astype(str).str.strip() == 'NIFTY']
    df = df[df['opt_type'].isin(['CE','PE'])]
    df['strike'] = pd.to_numeric(df['strike'], errors='coerce')
    df['open']   = pd.to_numeric(df['open'],   errors='coerce')
    return df[['expiry','strike','opt_type','open']].dropna(subset=['expiry','strike','open']).copy()

_bhav_mem = {}

def load_bhav(trade_date):
    if trade_date in _bhav_mem: return _bhav_mem[trade_date]
    cache_path = BHAV_CACHE / f'{trade_date}.pkl'
    if cache_path.exists():
        with open(cache_path, 'rb') as f: result = pickle.load(f)
        _bhav_mem[trade_date] = result; return result
    dd = trade_date.strftime('%d'); yyyy = trade_date.strftime('%Y')
    mon = trade_date.strftime('%b').upper(); yyyymmdd = trade_date.strftime('%Y%m%d')
    urls = [
        f'https://nsearchives.nseindia.com/content/fo/BhavCopy_FOAll_{yyyymmdd}_1.csv.zip',
        f'https://archives.nseindia.com/content/historical/DERIVATIVES/{yyyy}/{mon}/fo{dd}{mon}{yyyy}bhav.csv.zip',
    ]
    s = _get_nse_session(); result = None
    for url in urls:
        try:
            r = s.get(url, timeout=30)
            if r.status_code == 200 and len(r.content) > 1000:
                df = _parse_bhav_zip(r.content)
                if df is not None and len(df) > 0:
                    result = df; _time.sleep(0.25); break
        except: continue
    with open(cache_path, 'wb') as f: pickle.dump(result, f)
    _bhav_mem[trade_date] = result; return result

def get_option_open(trade_date, expiry, strike, opt_type):
    bhav = load_bhav(trade_date)
    if bhav is None: return None
    mask = ((bhav['expiry'] == expiry) &
            (bhav['strike'].round(0).astype(int) == int(strike)) &
            (bhav['opt_type'] == opt_type))
    rows = bhav[mask]
    if rows.empty: return None
    price = float(rows.iloc[0]['open'])
    return price if price > 0 else None

def _is_trading_day(d):
    if d.weekday() >= 5: return False
    if d in NSE_HOLIDAYS: return False
    if _trading_dates: return d in _trading_dates
    return True

def get_expiry(trade_date):
    days_ahead = (1 - trade_date.weekday()) % 7 if trade_date >= EXPIRY_CHANGE_DATE \
                 else (3 - trade_date.weekday()) % 7
    expiry = trade_date + timedelta(days=days_ahead)
    for _ in range(7):
        if _is_trading_day(expiry): break
        expiry -= timedelta(days=1)
    return expiry

def bs_price(S, K, T, r, sigma, opt_type='CE'):
    if T <= 1e-7: return max(0.0,S-K) if opt_type=='CE' else max(0.0,K-S)
    sq=sigma*np.sqrt(T); d1=(np.log(S/K)+(r+0.5*sigma**2)*T)/sq; d2=d1-sq
    if opt_type=='CE': return float(S*_norm.cdf(d1)-K*np.exp(-r*T)*_norm.cdf(d2))
    return float(K*np.exp(-r*T)*_norm.cdf(-d2)-S*_norm.cdf(-d1))

def implied_vol(S, K, T, r, market_price, opt_type):
    if T<=1e-7 or market_price<=0: return None
    intrinsic=max(0.0,S-K) if opt_type=='CE' else max(0.0,K-S)
    if market_price < intrinsic-1.0: return None
    try:
        sol=newton(lambda sig: bs_price(S,K,T,r,max(sig,1e-6),opt_type)-market_price, x0=0.20,tol=1e-5,maxiter=100)
        return float(sol) if 0.01<=sol<=5.0 else None
    except: return None

# ═══════════════════════════════════════════════════════════════════════════════
# 3. SIGNAL CLASSIFICATION (done once)
# ═══════════════════════════════════════════════════════════════════════════════
GAP_THR = 0.0015; GAP_LARGE = 0.0050

def compute_signals(row):
    return {
        'Gap Up'         : float(row['gap_pct']) >  GAP_THR,
        'Gap Up Strong'  : float(row['gap_pct']) >  GAP_LARGE,
        'Gap Down'       : float(row['gap_pct']) < -GAP_THR,
        'Prev India UP'  : float(row['prev_india_ret']) > 0,
        'Prev India DOWN': float(row['prev_india_ret']) < 0,
        'US UP'          : float(row['SP500_ret']) > 0,
        'US DOWN'        : float(row['SP500_ret']) < 0,
        'SGX UP'         : float(row['SGX_ret']) > 0,
        'SGX DOWN'       : float(row['SGX_ret']) < 0,
        'DAX UP'         : float(row['DAX_ret']) > 0,
        'VIX Rising'     : float(row['VIX_US_ret']) > 0.03,
        'VIX Falling'    : float(row['VIX_US_ret']) < 0,
        'VIX Spike'      : float(row['VIX_US_ret']) > 0.05,
    }

def check_combo(signals, signal_str):
    return all(signals.get(s.strip(), False) for s in signal_str.split('+'))

bear_combos = list(top_bearish['Signal'])

print('\nClassifying signals...')
tradeable_rows = []
for _, row in backtest_df.iterrows():
    d = row['india_date'].date()
    if d.weekday() == 0: continue
    if d in EVENT_DAYS: continue
    sigs = compute_signals(row)
    bear_fired = [c for c in bear_combos if check_combo(sigs, c)]
    if not bear_fired: continue
    tradeable_rows.append({
        'india_date': row['india_date'],
        'india_open': float(row['india_open']),
        'vix_india' : float(row['VIX_INDIA_level']),
        'dir_60'    : int(row['dir_60']),
        'ret_60'    : float(row['ret_60']),
        'combo'     : bear_fired[0],
    })

tradeable_df = pd.DataFrame(tradeable_rows)
print(f'Tradeable days (BEARISH): {len(tradeable_df)}')

# ═══════════════════════════════════════════════════════════════════════════════
# 4. SIMULATE TRADES — cached per (SL, TP)
#    Entry price & IV are the same for a given day regardless of SL/TP.
#    So we compute entry/IV once, then apply SL/TP logic separately.
# ═══════════════════════════════════════════════════════════════════════════════

print('\nPre-fetching bhav copies & computing entry prices / IV for all tradeable days...')
entry_cache = {}   # trade_date -> {expiry, dte, opt, strike, atm, entry_price, iv, day_minute_bars}

for _, trow in tradeable_df.iterrows():
    d = trow['india_date'].date()
    if d in entry_cache: continue

    expiry = get_expiry(d)
    dte    = (expiry - d).days
    nifty_open = trow['india_open']
    atm    = round(nifty_open / STRIKE_STEP) * STRIKE_STEP
    strike = atm - STRIKE_STEP * STRIKES_OTM
    opt    = 'PE'

    entry_price = get_option_open(d, expiry, int(strike), opt)
    if entry_price is None:
        entry_cache[d] = None
        print(f'  {d}: SKIPPED (no bhav)')
        continue

    expiry_dt = datetime.combine(expiry, dtime(15, 30))
    def tte(t_obj, _d=d, _exp_dt=expiry_dt):
        return max((
            _exp_dt - datetime.combine(_d, t_obj)
        ).total_seconds(), 0.0) / (365.25 * 24 * 3600)

    iv = implied_vol(nifty_open, strike, tte(dtime(9, 15)), RISK_FREE_RATE, entry_price, opt)
    if iv is None:
        iv_raw = trow['vix_india'] / 100.0
        iv = iv_raw * (1.30 if dte <= 2 else 1.15 if dte <= 4 else 1.05)

    day_min = (minute_all[minute_all.index.date == d].between_time('09:16', '10:15')
               if not minute_all.empty else pd.DataFrame())

    entry_cache[d] = {
        'expiry': expiry, 'dte': dte, 'opt': opt,
        'strike': strike, 'atm': atm,
        'entry_price': entry_price, 'iv': iv,
        'day_min': day_min, 'tte': tte,
    }
    print(f'  {d}: entry={entry_price:.1f}  IV={iv*100:.1f}%  DTE={dte}')

# ─────────────────────────────────────────────────────────────────────────────
# Simulate one (SL, TP) sweep using cached entry data
# ─────────────────────────────────────────────────────────────────────────────
def simulate_sl_tp(sl_pct, tp_pct):
    """Return list of trade dicts for given SL/TP using cached entry data."""
    trades = []
    for _, trow in tradeable_df.iterrows():
        d = trow['india_date'].date()
        ec = entry_cache.get(d)
        if ec is None: continue

        ep      = ec['entry_price']
        iv      = ec['iv']
        strike  = ec['strike']
        opt     = ec['opt']
        tte_fn  = ec['tte']
        day_min = ec['day_min']

        sl_price = ep * (1 - sl_pct)
        tp_price = ep * (1 + tp_pct)

        exit_price  = None
        exit_reason = '10:15 exit'
        exit_time   = '10:15'

        for ts, m in day_min.iterrows():
            spot   = float(m['close'])
            t_now  = ts.to_pydatetime().replace(tzinfo=None).time()
            price  = bs_price(spot, strike, tte_fn(t_now), RISK_FREE_RATE, iv, opt)
            sl_active = t_now >= dtime(9, 20)  # SL locked out for first 5 min (9:15-9:19)
            if sl_active and price <= sl_price:
                exit_price, exit_reason, exit_time = sl_price, 'Stop Loss', str(t_now)[:5]; break
            if price >= tp_price:
                exit_price, exit_reason, exit_time = tp_price, 'Target Hit', str(t_now)[:5]; break

        if exit_price is None:
            spot_last = float(day_min.iloc[-1]['close']) if len(day_min) > 0 else trow['india_open']
            exit_price = bs_price(spot_last, strike, tte_fn(dtime(10, 15)), RISK_FREE_RATE, iv, opt)

        pnl_pts = round(exit_price - ep, 2)
        pnl_rs  = round(pnl_pts * LOT_SIZE - BROKERAGE, 2)

        trades.append({
            'Date'       : d,
            'Combo'      : trow['combo'],
            'NIFTY Open' : int(round(trow['india_open'])),
            'Strike'     : ec['strike'],
            'ATM'        : ec['atm'],
            'Expiry'     : ec['expiry'],
            'DTE'        : ec['dte'],
            'IV%'        : round(iv * 100, 1),
            'Entry (pts)': round(ep, 2),
            'Exit (pts)' : round(exit_price, 2),
            'Exit Reason': exit_reason,
            'Exit Time'  : exit_time,
            'P&L (pts)'  : pnl_pts,
            'P&L (Rs)'   : pnl_rs,
            'Correct?'   : 'YES' if trow['dir_60'] == -1 else 'NO',
            'Actual'     : 'DOWN' if trow['dir_60'] == -1 else 'UP',
            'Actual Ret%': f"{trow['ret_60']:+.2%}",
        })
    return trades

# ═══════════════════════════════════════════════════════════════════════════════
# 5. CAPITAL METRICS
# ═══════════════════════════════════════════════════════════════════════════════
def compute_xirr(cf, dates):
    try:
        return newton(
            lambda r: sum(c / ((1+r) ** ((d - dates[0]).days / 365.0)) for c, d in zip(cf, dates)),
            0.1)
    except: return None

def capital_metrics(trades_list, fixed, max_lots, dte0_max_lots):
    """Compute capital curve metrics for given capital params."""
    if not trades_list:
        return {}

    capital = None
    ic      = 0.0
    tr_refill = 0.0
    rc      = 0
    cf      = []
    cd      = []
    cap_curve = []
    pnl_list  = []

    for row in trades_list:
        ep  = float(row['Entry (pts)'])
        pp  = float(row['P&L (pts)'])
        cpl = ep * LOT_SIZE
        if cpl <= 0: continue

        td  = pd.to_datetime(row['Date'])

        if capital is None:
            capital = cpl * fixed
            ic      = capital
            cf.append(-ic); cd.append(td)

        cb  = capital
        req = cpl * fixed
        ra  = 0.0
        if cb < req:
            ra  = req - cb
            cb += ra
            capital = cb
            rc += 1
            tr_refill += ra
            cf.append(-ra); cd.append(td)

        car = cb
        unc = max(fixed, int(car / cpl))
        dv  = int(row['DTE'])
        d0  = min(unc, dte0_max_lots) if dv == 0 else unc
        lots = min(d0, max_lots)

        tp  = pp * LOT_SIZE * lots - BROKERAGE * lots
        capital = car + tp
        cap_curve.append(round(capital, 2))
        pnl_list.append(tp)

    if not cap_curve:
        return {}

    cf.append(capital); cd.append(pd.to_datetime(trades_list[-1]['Date']))
    ti   = ic + tr_refill
    net  = (capital - ti) / ti * 100 if ti > 0 else 0
    xirr = compute_xirr(cf, cd)

    total_tp = sum(pnl_list)
    wins     = sum(1 for p in pnl_list if p > 0)
    n        = len(pnl_list)
    avg_pnl  = total_tp / n if n > 0 else 0

    # Drawdown
    peak = ic
    max_dd = 0.0
    running = ic
    for p in pnl_list:
        running += p
        if running > peak: peak = running
        dd = (peak - running) / peak * 100
        if dd > max_dd: max_dd = dd

    return {
        'Trades'          : n,
        'Wins'            : wins,
        'Win Rate %'      : round(wins / n * 100, 1) if n else 0,
        'Total PnL (Rs)'  : round(total_tp, 0),
        'Avg PnL/Trade (Rs)': round(avg_pnl, 0),
        'Initial Capital' : round(ic, 0),
        'Final Capital'   : round(capital, 0),
        'Total Invested'  : round(ti, 0),
        'Net Return %'    : round(net, 2),
        'XIRR %'          : round(xirr * 100, 1) if xirr else None,
        'Refill Count'    : rc,
        'Total Refilled'  : round(tr_refill, 0),
        'Max Drawdown %'  : round(max_dd, 2),
    }

# ═══════════════════════════════════════════════════════════════════════════════
# 6. RUN FULL GRID
# ═══════════════════════════════════════════════════════════════════════════════
# (MAX_LOTS, DTE0) valid pairs
valid_lot_pairs = [
    (ml, d0)
    for ml in MAX_LOTS_GRID
    for d0 in DTE0_LOTS_GRID
    if d0 <= ml
]

# Pre-compute trades for each (SL, TP) pair
sl_tp_cache = {}
sl_tp_combos = list(itertools.product(SL_GRID, TP_GRID))

print(f'\nRunning {len(sl_tp_combos)} (SL,TP) combos × {len(valid_lot_pairs)} lot configs × {len(FIXED_GRID)} FIXED values')
print(f'= {len(sl_tp_combos) * len(valid_lot_pairs) * len(FIXED_GRID)} total combinations\n')

all_rows = []
combo_idx = 0
total_combos = len(sl_tp_combos) * len(valid_lot_pairs) * len(FIXED_GRID)

for sl, tp in sl_tp_combos:
    key = (sl, tp)
    if key not in sl_tp_cache:
        print(f'  Simulating trades: SL={sl:.0%} TP={tp:.0%}...', end='  ', flush=True)
        sl_tp_cache[key] = simulate_sl_tp(sl, tp)
        trades = sl_tp_cache[key]
        n = len(trades)
        wr = sum(1 for t in trades if t['P&L (Rs)'] > 0) / n * 100 if n else 0
        tot = sum(t['P&L (Rs)'] for t in trades)
        print(f'{n} trades  WR={wr:.1f}%  1-lot PnL=Rs{tot:,.0f}')

    trades = sl_tp_cache[key]

    for fixed in FIXED_GRID:
        for max_lots, dte0 in valid_lot_pairs:
            combo_idx += 1
            m = capital_metrics(trades, fixed, max_lots, dte0)
            if not m: continue
            row = {
                'Combo #'         : combo_idx,
                'SL %'            : f'{sl:.0%}',
                'TP %'            : f'{tp:.0%}',
                'SL_val'          : sl,
                'TP_val'          : tp,
                'FIXED'           : fixed,
                'MAX_LOTS'        : max_lots,
                'DTE0_MAX_LOTS'   : dte0,
                **m,
            }
            all_rows.append(row)

            if combo_idx % 20 == 0:
                print(f'  Progress: {combo_idx}/{total_combos}')

summary_df = pd.DataFrame(all_rows)
print(f'\nGrid search complete. {len(summary_df)} result rows.')

# ═══════════════════════════════════════════════════════════════════════════════
# 7. WRITE EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

output_path = OUTPUT_DIR / 'grid_search_results.xlsx'

# Sort by XIRR desc as default view
summary_df_sorted = summary_df.sort_values('XIRR %', ascending=False).reset_index(drop=True)

# Build per-SL/TP trade log sheets
trade_log_sheets = {}
for (sl, tp), trades in sl_tp_cache.items():
    if trades:
        trade_log_sheets[f'Trades_SL{int(sl*100)}_TP{int(tp*100)}'] = pd.DataFrame(trades)

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # ── Sheet 1: Master Summary ──
    display_cols = [
        'SL %', 'TP %', 'FIXED', 'MAX_LOTS', 'DTE0_MAX_LOTS',
        'Trades', 'Wins', 'Win Rate %', 'Avg PnL/Trade (Rs)',
        'Total PnL (Rs)', 'Initial Capital', 'Final Capital',
        'Net Return %', 'XIRR %', 'Max Drawdown %',
        'Refill Count', 'Total Refilled',
    ]
    summary_df_sorted[display_cols].to_excel(writer, sheet_name='Master Summary', index=False)

    # ── Sheet 2: Top 20 by XIRR ──
    top20 = summary_df_sorted.head(20)[display_cols]
    top20.to_excel(writer, sheet_name='Top 20 by XIRR', index=False)

    # ── Sheet 3: Top 20 by Net Return % ──
    top20_net = summary_df.sort_values('Net Return %', ascending=False).head(20)[display_cols]
    top20_net.to_excel(writer, sheet_name='Top 20 by Net Return', index=False)

    # ── Sheet 4: Pivot — XIRR by SL vs TP (FIXED=5, MAX_LOTS=100, DTE0=100) ──
    for fixed_val in FIXED_GRID:
        pivot_df = (summary_df[(summary_df['FIXED'] == fixed_val) &
                               (summary_df['MAX_LOTS'] == 100) &
                               (summary_df['DTE0_MAX_LOTS'] == 100)]
                    .pivot_table(index='SL %', columns='TP %', values='XIRR %', aggfunc='first'))
        pivot_df.to_excel(writer, sheet_name=f'XIRR Pivot FIXED={fixed_val}')

    # ── Sheet 5: XIRR by FIXED (SL=0.20, TP=0.40) ──
    lot_pivot = (summary_df[(summary_df['SL_val'] == 0.20) & (summary_df['TP_val'] == 0.40)]
                 .pivot_table(index=['MAX_LOTS', 'DTE0_MAX_LOTS'], columns='FIXED',
                              values='XIRR %', aggfunc='first'))
    lot_pivot.to_excel(writer, sheet_name='XIRR vs Lot Config')

    # ── Trade logs ──
    for sheet_name, df_log in list(trade_log_sheets.items())[:8]:  # cap at 8 to avoid huge file
        df_log.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    # ── Config sheet ──
    pd.DataFrame([
        {'Parameter': 'Signal Mode',       'Value': SIGNAL_MODE},
        {'Parameter': 'Backtest Start',    'Value': str(start_date)},
        {'Parameter': 'Backtest End',      'Value': str(end_date)},
        {'Parameter': 'Total Days',        'Value': len(backtest_df)},
        {'Parameter': 'Tradeable Days',    'Value': len(tradeable_df)},
        {'Parameter': 'SL Grid',           'Value': str(SL_GRID)},
        {'Parameter': 'TP Grid',           'Value': str(TP_GRID)},
        {'Parameter': 'FIXED Grid',        'Value': str(FIXED_GRID)},
        {'Parameter': 'MAX_LOTS Grid',     'Value': str(MAX_LOTS_GRID)},
        {'Parameter': 'DTE0 Grid',         'Value': str(DTE0_LOTS_GRID)},
        {'Parameter': 'LOT_SIZE',          'Value': LOT_SIZE},
        {'Parameter': 'BROKERAGE/lot',     'Value': BROKERAGE},
        {'Parameter': 'STRIKES_OTM',       'Value': STRIKES_OTM},
        {'Parameter': 'Entry Source',      'Value': 'NSE F&O Bhav Copy (real OPEN)'},
        {'Parameter': 'Exit Method',       'Value': 'NIFTY spot minute + real IV (BS path)'},
        {'Parameter': 'Monday Filter',     'Value': 'ON'},
        {'Parameter': 'Event Filter',      'Value': f'ON ({len(EVENT_DAYS)} days)'},
        {'Parameter': 'Total Combinations','Value': len(summary_df)},
    ]).to_excel(writer, sheet_name='Config', index=False)

# ── Post-process: formatting ──────────────────────────────────────────────────
wb = load_workbook(output_path)

HEADER_FILL = PatternFill('solid', start_color='1F3864', end_color='1F3864')
ALT_FILL    = PatternFill('solid', start_color='EBF3FB', end_color='EBF3FB')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT   = Font(name='Arial', size=9)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left',   vertical='center')
thin        = Side(style='thin', color='BFBFBF')
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws, num_format_map=None):
    """Apply consistent styling to a worksheet."""
    if num_format_map is None: num_format_map = {}
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(cell.value or '')) + 4, 12)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = BORDER
            col_name       = ws.cell(1, cell.column).value or ''
            if col_name in num_format_map:
                cell.number_format = num_format_map[col_name]
            elif 'Rs' in str(col_name) or 'Capital' in str(col_name) or 'PnL' in str(col_name) or 'Refill' in str(col_name):
                cell.number_format = '#,##0'
                cell.alignment     = CENTER
            elif '%' in str(col_name):
                cell.number_format = '0.0'
                cell.alignment     = CENTER
            else:
                cell.alignment     = CENTER

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

# Style all main sheets
for sheet_name in ['Master Summary', 'Top 20 by XIRR', 'Top 20 by Net Return']:
    if sheet_name in wb.sheetnames:
        style_sheet(wb[sheet_name])

# Color scale on XIRR % column in Master Summary
ws_main = wb['Master Summary']
xirr_col = None
for cell in ws_main[1]:
    if cell.value == 'XIRR %':
        xirr_col = cell.column_letter
        break
if xirr_col:
    last_row = ws_main.max_row
    ws_main.conditional_formatting.add(
        f'{xirr_col}2:{xirr_col}{last_row}',
        ColorScaleRule(start_type='min', start_color='FF6B6B',
                       mid_type='percentile', mid_value=50, mid_color='FFEB84',
                       end_type='max', end_color='63BE7B'))

# Style config sheet
if 'Config' in wb.sheetnames:
    ws_cfg = wb['Config']
    for cell in ws_cfg[1]:
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER
    ws_cfg.column_dimensions['A'].width = 22
    ws_cfg.column_dimensions['B'].width = 40

wb.save(output_path)
print(f'\n✓ Excel saved: {output_path}')

# ── Quick console summary ──────────────────────────────────────────────────────
print('\n' + '='*70)
print('TOP 10 CONFIGURATIONS BY XIRR %')
print('='*70)
top10 = summary_df_sorted.head(10)[['SL %','TP %','FIXED','MAX_LOTS','DTE0_MAX_LOTS',
                                      'Trades','Win Rate %','XIRR %','Net Return %','Max Drawdown %']]
print(top10.to_string(index=False))

print('\n' + '='*70)
print('TOP 10 CONFIGURATIONS BY NET RETURN %')
print('='*70)
top10_net = summary_df.sort_values('Net Return %', ascending=False).head(10)[
    ['SL %','TP %','FIXED','MAX_LOTS','DTE0_MAX_LOTS','Trades','Win Rate %','XIRR %','Net Return %','Max Drawdown %']]
print(top10_net.to_string(index=False))

print(f'\nDone. {len(summary_df)} combinations evaluated.')
print(f'Output: {output_path}')